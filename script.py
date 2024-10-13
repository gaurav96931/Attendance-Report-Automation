import pandas as pd
from datetime import datetime, time
# openpyxl is library used to work with excel sheets in python
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_FILE = 'output.xlsx'
START_TIME = time(18)
END_TIME = time(20)

# Steps:
# 1) parse takes input of a string having date and time info of a record
# 2) strptime converts timestamp into datetime object (here date_time)
# 3) extract date & time from date_time, convert date into required format using strftime, return both
def parse(timestamp):
    date_time = datetime.strptime(timestamp, '%d/%m/%Y %H:%M:%S') 
    date = datetime.date(date_time).strftime('%d/%m/%Y') 
    time = date_time.time()
    return date, time

# Function to parse date strings
def date_key(date_string):
    return datetime.strptime(date_string, '%d/%m/%Y')

# using try except block to check if string value can be type casted to int
def is_int(x):
    try:
        x = int(x)
        return True
    except:
        return False

# class dates data
classes_taken_dates = ["06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024", "03/09/2024", "17/09/2024", "01/10/2024"]
classes_missed_dates = ["10/09/2024"]
exams_dates = ["24/09/2024"]

# to store all the above listed dates
concerned_dates = []

for date in classes_taken_dates:
    concerned_dates.append(date)
for date in classes_missed_dates:
    concerned_dates.append(date)
for date in exams_dates:
    concerned_dates.append(date)

concerned_dates.sort(key=date_key)

# taking input of students enrolled in the course
students = []
with open('stud_list.txt', 'r') as f:
    students = [line.rstrip() for line in f]

# to store data of students' attendance in dictionary of dictionary format
excel_data = {}
for date in concerned_dates:
    excel_data[date] = {}
excel_data["Total count of dates"]={}
excel_data["Total Attendance Marked"]={}
excel_data["Total Attendance allowed"]={}
excel_data["Proxy"]={}

for date in concerned_dates:
    for student in students:
        excel_data[date][student] = 0
for student in students:
    excel_data['Proxy'][student] = 0
    excel_data["Total Attendance Marked"][student] = 0
    excel_data["Total count of dates"][student] = 0
    excel_data["Total Attendance allowed"][student]=len(classes_taken_dates)*2

attendance_data = pd.read_csv('input_attendance.csv', header=0, parse_dates=[0])

for index, row in attendance_data.iterrows():
    try:
        record_date, record_time = parse(row["Timestamp"])
        student = row["Roll"]
        if record_date in concerned_dates:
            excel_data[record_date][student] += 1
            excel_data["Total count of dates"][student] += 1
        else:
            excel_data['Proxy'][student] += 1
        excel_data["Total Attendance Marked"][student] += 1
        
    except:
        print(f"Error on reading line {index+2} in input_attendance.csv")


# saving the data
df = pd.DataFrame(excel_data)
df.to_excel(OUTPUT_FILE)

# formatting the excel sheet
wb = load_workbook(OUTPUT_FILE)
ws = wb.active
ws['A1'] = 'Roll'
# Create default style objects
default_font = Font()  # Default font
default_fill = PatternFill(fill_type=None)  # No fill
default_alignment = Alignment()  # Default alignment
default_border = Border(left=Side(style=None), right=Side(style=None), 
                        top=Side(style=None), bottom=Side(style=None))  # No border

# Loop through all cells and reset formatting 
for row in ws.iter_rows():
    for cell in row:
        cell.font = default_font
        cell.alignment = default_alignment
        cell.border = default_border
        cell.fill = default_fill
        if cell.value in classes_missed_dates:  # highlight red for all missed classes dates
            cell.fill = PatternFill(fill_type="solid", start_color="FF474C", end_color="FF474C")

# highlighting according to cell value
for row in ws.iter_rows(min_col=1, max_col=len(concerned_dates)):
    for cell in row:
        if is_int(cell.value):
            if cell.value == 2:
                cell.fill = PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00")
            elif cell.value == 1:
                cell.fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
            elif int(cell.value) > 2:
                cell.fill = PatternFill(fill_type="solid", start_color="FF474C", end_color="FF474C")

# saving finalized workbook
wb.save(OUTPUT_FILE)